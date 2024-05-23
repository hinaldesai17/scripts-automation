def download_excel_from_ice(date_str):
    import requests
    
    '''Function to download the excel file from ICE.com'''

    url = "https://www.ice.com/marketdata/publicdocs/futures_us_reports/cocoa/cocoa_cert_stock_" + date_str + ".xls"

    response = requests.get(url)
    filename = "cocoa_cert_stock_" + date_str + ".xls"

    print("URL: ", url)
    if response.status_code == 200:
        with open(filename, "wb") as file:
            file.write(response.content)
            print("File downloaded successfully")
            return filename
    else:
        print("Failed to download file")
        return None

def get_worksheet(filename, date_str):
    import pyexcel as p
    import openpyxl
    
    print("Hii:", f"{date_str}.xlsx")
    p.save_book_as(file_name=filename, dest_file_name=f"{date_str}.xlsx")


    wb = openpyxl.load_workbook(f"{date_str}.xlsx")

    print("Converted to xlsx format")

    ws = wb.active #first worksheet

    print("Rows: ", str(ws.max_row))
    print("Columns: ", str(ws.max_column))
    return ws

def get_date_from_sheet(ws):
    from datetime import datetime
    #A1 contains date
    date_str = ws['A1'].value
    date_part = date_str.replace("Date: ", "")
    # Parse the date string into a datetime object
    date = datetime.strptime(date_part, "%m/%d/%Y")
    print(date)
    return date

def prepare_df_from_excel(ws):
    import pandas as pd
    
    data = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True):
        data.append(row)

    df = pd.DataFrame(data)
    return df

def prepare_cocoa_cert_stk_df(df, date):
    # Find the index of the row containing "Cocoa Certified Stock by Port & Growth - Bags"
    index_of_start_row = df[df[0] == "Cocoa Certified Stock by Port & Growth - Bags"].index[0]

    # Find the index of the row containing "Cocoa Grading Results in Bags"
    index_of_end_row = df[df[0].str.contains("Cocoa Grading Results in Bags", na=False)].index[0]

    # Slice the DataFrame from the row after "Cocoa Certified Stock by Port & Growth - Bags" 
    # to the row before "Cocoa Grading Results in Bags"
    cocoa_cert_stk_df = df.iloc[index_of_start_row + 1:index_of_end_row].reset_index(drop=True)

    # Drop rows and columns with all NaN values
    cocoa_cert_stk_df = cocoa_cert_stk_df.dropna(how='all')
    cocoa_cert_stk_df = cocoa_cert_stk_df.dropna(how='all', axis=1)

    cocoa_cert_stk_df.iloc[0] = cocoa_cert_stk_df.iloc[0].fillna("SYM")

    index_to_drop = cocoa_cert_stk_df.columns[cocoa_cert_stk_df.iloc[0].str.startswith('SYM')].tolist()[0]
    # Drop the entire column
    cocoa_cert_stk_df = cocoa_cert_stk_df.drop(cocoa_cert_stk_df.columns[index_to_drop], axis=1)

    cocoa_cert_stk_df = cocoa_cert_stk_df.rename(columns=cocoa_cert_stk_df.iloc[0])
    cocoa_cert_stk_df = cocoa_cert_stk_df.drop(1)
    cocoa_cert_stk_df.set_index("SYM", inplace=True)

    # Create a dictionary to keep track of column name occurrences
    column_counts = {}

    # Rename columns based on their occurrences
    new_columns = []
    for column in cocoa_cert_stk_df.columns:
        if column in column_counts:
            new_name = f"{column}_lots"
            column_counts[column] += 1
        else:
            new_name = f"{column}_bags"
            column_counts[column] = 1
        new_columns.append(new_name)

    cocoa_cert_stk_df.columns = new_columns
    cocoa_cert_stk_df.reset_index(inplace=True)

    cocoa_cert_stk_df['date'] = date

    cocoa_cert_stk_df_bags = cocoa_cert_stk_df[[col for col in cocoa_cert_stk_df.columns if col.endswith('_bags')]]
    new_columns = [col.split('_')[0] for col in cocoa_cert_stk_df_bags.columns]
    cocoa_cert_stk_df_bags.columns = new_columns
    cocoa_cert_stk_df_bags['date'] = date
    cocoa_cert_stk_df_bags['SYM'] = cocoa_cert_stk_df['SYM']

    cocoa_cert_stk_df_lots = cocoa_cert_stk_df[[col for col in cocoa_cert_stk_df.columns if col.endswith('_lots')]]
    new_columns = [col.split('_')[0] for col in cocoa_cert_stk_df_lots.columns]
    cocoa_cert_stk_df_lots.columns = new_columns
    cocoa_cert_stk_df_lots['date'] = date
    cocoa_cert_stk_df_lots['SYM'] = cocoa_cert_stk_df['SYM']

    cocoa_cert_stk_df_bags.dropna(inplace=True)
    cocoa_cert_stk_df_lots.dropna(inplace=True)
    cocoa_cert_stk_df_bags.loc[cocoa_cert_stk_df_bags['SYM'] == 'Total Lots', 'SYM'] = 'Total Bags'

    return cocoa_cert_stk_df_bags, cocoa_cert_stk_df_lots


def prepare_cocoa_grading_results_df(df, date):
    index_of_start_row = df[df[0].str.contains("Cocoa Grading Results in Bags", na=False)].index[0]
    if(df.iloc[index_of_start_row+1][0] == "No Grading to Date"):
        return None, None

    index_of_end_row = df[df[0].str.contains("Total Bags Reported By ICE FUTURES U.S. Licensed Warehouses", na=False)].index[0]

    cocoa_grading_results_df = df.iloc[index_of_start_row + 1:index_of_end_row].reset_index(drop=True)

    cocoa_grading_results_df = cocoa_grading_results_df.dropna(how='all')
    cocoa_grading_results_df = cocoa_grading_results_df.dropna(how='all', axis=1)

    cocoa_grading_results_df.columns = ['SYM', 'pending_grad_bags', 'passed_grad_bags', 'SYM2', 'pending_grad_lots', 'passed_grad_lots']
    cocoa_grading_results_df.drop(1, inplace=True)
    cocoa_grading_results_df.drop(2, inplace=True)
    cocoa_grading_results_df.drop('SYM2', axis=1, inplace=True)
    cocoa_grading_results_df.reset_index(drop=True, inplace=True)
    cocoa_grading_results_df['date'] = date

    if cocoa_grading_results_df.empty:
        return None, None

    cocoa_grading_results_df_bags = cocoa_grading_results_df[[col for col in cocoa_grading_results_df.columns if col.endswith('_bags')]]
    cocoa_grading_results_df_bags['date'] = date
    cocoa_grading_results_df_bags['SYM'] = cocoa_grading_results_df['SYM']

    cocoa_grading_results_df_lots = cocoa_grading_results_df[[col for col in cocoa_grading_results_df.columns if col.endswith('_lots')]]
    cocoa_grading_results_df_lots['date'] = date
    cocoa_grading_results_df_lots['SYM'] = cocoa_grading_results_df['SYM']

    cocoa_grading_results_df_bags.columns = ['pending_grad', 'passed_grad', 'date', 'SYM']
    cocoa_grading_results_df_lots.columns = ['pending_grad', 'passed_grad', 'date', 'SYM']

    return cocoa_grading_results_df_bags, cocoa_grading_results_df_lots

def prepare_cocoa_total_bags_df(df, date):
    index_of_start_row = df[df[0].str.contains("Total Bags Reported By ICE FUTURES U.S. Licensed Warehouses", na=False)].index[0]

    index_of_end_row = df[df[0].str.contains("GRAND TOTAL:", na=False)].index[0]

    cocoa_total_bags_df = df.iloc[index_of_start_row+1:index_of_end_row + 1].reset_index(drop=True)

    cocoa_total_bags_df = cocoa_total_bags_df.dropna(how='all')
    cocoa_total_bags_df = cocoa_total_bags_df.dropna(how='all', axis=1)
    cocoa_total_bags_df.columns = ['SYM', 'value']
    cocoa_total_bags_df.reset_index(drop=True, inplace=True)

    cocoa_total_bags_df['date'] = date
    return cocoa_total_bags_df